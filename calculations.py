from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


def invoice_generator(filename):
    wb=load_workbook(filename)
    sheet=wb.create_sheet('Invoice Generator')
    sheet.append(['Product','Price/unit','Quantiy','Amount'])
    
    print(f'{"\nget your invoice!!".title():=>37}\n{"="*18}\n')

    num_prod=int(input('Enter the number of the products you bought: '))
    print('\nNow Enter the product details:\n')
    total=0
    for prod in range(num_prod):
        name=input(f'Product {prod+1}: ')
        price=float(input('Price: '))
        quantity=int(input('Quantity: '))
        amount=price*quantity
        total+=amount
        sheet.append((name,price,quantity,amount))
        
    cell_total=sheet.cell(sheet.max_row+1,1)
    cell_total.value='Total'
    cell_total=sheet.cell(sheet.max_row,sheet.max_column)
    cell_total.value=total
    wb.save(filename)
    wb.close()
      
            
def process_workbook(filename):
    wb=load_workbook(filename)
    sheet= wb['Summary']
 
    percent_cell=sheet.cell(3,6)
    percent_cell.value='Percentage Profit'
    amount_column=4
    for row in range(4, sheet.max_row+1):
        cell_amount=sheet.cell(row,amount_column) 
        correct_cell_amount=sheet.cell(row,6)
        correct_cell_amount.value=f'%{abs(100-(cell_amount.value // 5))}'

    values=Reference(sheet, 
                    min_row=4, 
                    max_row=sheet.max_row, 
                    min_col=4, 
                    max_col=4)
    amount_chart= BarChart()
    amount_chart.add_data(values)
    sheet.add_chart(amount_chart,'h3')

    wb.save(filename)
    wb.close()




























